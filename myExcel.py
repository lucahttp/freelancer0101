from time import sleep
import datetime

#SELENIUM
# import pytest
import time
import json

#EXCEL
from openpyxl import load_workbook
import os
import sys
import openpyxl #Connect the library
from openpyxl import Workbook
from openpyxl.styles import PatternFill#Connect cell styles
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill#Connect styles for text
from openpyxl.styles import colors#Connect colors for text and cells
from openpyxl.utils import get_column_letter
import datetime
import string


# local
import myMail
import myWord
import myQRCode
import myConfig

CurrentWorksheet = ""


###TOOLS CROSS SCRIPT

def GetRandom():
    #https://docs.python.org/3/library/random.html
    import random
    random_number = random.randrange(0, 10000, 1)
    print(random_number)
    #excelFile = "tempFile-{}.xlsx".format(str(current_time))
    #excelFile = "tempFile-{}.xlsx".format(random_number)
    return random_number

def getTime():
    import datetime
    x = datetime.datetime.now()
    xs = x.strftime("%X %x")
    datestring = str(xs)
    datestring = datestring.replace(":","-")
    datestring = datestring.replace("/", "-")
    datestring = datestring.replace(" ", "--")
    print(datestring)
    return datestring


###TOOLS CROSS SCRIPT


def getDateNormaliced(data_raw):
    from datetime import datetime

    if type(data_raw) == datetime:
        #date_str = '09-19-2018'
        #date_object = datetime.strptime(data_raw, '%d-%m-%Y').date()
        #print(type(date_object))
        #print(date_object)  # printed in default formatting
        date_formated = data_raw.strftime('%d/%m/%Y')
        print(date_formated)
        return date_formated
    else:
        print("invalid date")
        #return "invalid date"
        return data_raw
    pass
### SUB-SCRIPTS
#  

import os

def CreateExample():
    ##########################################################
    ##  Chapter One Creating the excel to fill data
    ##########################################################


    #https://stackabuse.com/the-python-tempfile-module/


    tempfilename = "~example.xlsx"
    tempfilename = "example.xlsx"


    datestring = getTime()
    #with io.open((path.replace("xlsx", "{}.json")).replace("xls", "{}.json").format(number,number),'w',encoding='utf8') as f:

    ##tempfilename = "~tempfile{}.xlsx".format(datestring)

    
    #delete file if exist feature

    if os.path.exists(tempfilename):
        os.remove(tempfilename)
        print("The file was deleted")
    else:
        print("The file does not exist")
    


    #Styles
    tabColor_text = Font(size=11, underline='none', color = 'ffffffff', bold=False, italic=False) #what color = colors.RED — color prescribed in styles
    tabColor_cell = PatternFill(fill_type='solid', start_color='00000000', end_color='00000000')#This code allows you to do design color cells


    wb = Workbook()
    page = wb.active
    #Set the name of the workbook
    page.title = 'TO_SERVICE-NOW'

    #Insert the example data to the example excel    
    
    
    #First Name	Last Name	Email	Cell  Phone	Unit	Arrival	Departure	Code	Company Name	Contact	Maddress	Mcity	Mstate	Mzip	MEmail	Phone	Greeting	Salutation	LINK to QR
    #"First Name","Last Name","Email","Cell  Phone","Unit","Arrival","Departure","Code","Company Name","Contact","Maddress","Mcity","Mstate","Mzip","MEmail","Phone","Greeting","Salutation","LINK to QR"
    #"Randy","Ward","rward@lapro.tv","15042280136","0207A","04/09/2020","08/09/2020","0210207A24825215042280136Jamie Albert","Pam Martin - Keller Williams  ","Pam Martin"," REALTOR","18025 Lake Iris Avenue","Baton Rouge","LA","70817","pammartin01@yahoo.com pammartin01@kw.com","(251)269-8864 or (251)279-0716","Are you ready for your vacation?  ","We hope you enjoy your stay. Please contact me If you need anything!"
    #"Vicke","Ward","vrward@att.net","15042280137","0807A","14/09/2020","18/09/2020","0210207A24825215042280136Jamie Albert","Pam Martin - Keller Williams  ","Pam Martin"," REALTOR","18026 Lake Iris Avenue","Baton Rouge","LA","70817","pammartin01@yahoo.com pammartin01@kw.com","(251)269-8864 or (251)279-0717","Are you ready for your vacation?  ","We hope you enjoy your stay. Please contact me If you need anything!"
            
    # TabTitles = ["Assignment Group", "Free Text Field", "Done", "REQUEST", "RITM", "TASK", "DATE", "GROUP"]
    TabTitles = ["First Name","Last Name","Email","Cell  Phone","Unit","Arrival","Departure","Code","Company Name","Contact","Maddress","Mcity","Mstate","Mzip","MEmail","Phone","Greeting","Salutation","LINK to QR"]
    page.append(TabTitles) # write the TabTitles to the first line

    #Insert the example data to the example excel
    #companies = ["datavision","excuse me, this is a test","<YES/NO>","REQ0000000","RITM0000000","SCTASK0000000","aaaa-aa-aa a:aa:aa","BI-ITMS-DATAVISION-ACCT-MGMT"]
    companies = ["Randy","Ward","rward@lapro.tv","15042280136","0207A","04/09/2020","08/09/2020","0210207A24825215042280136Jamie Albert","Pam Martin - Keller Williams  ","Pam Martin, REALTOR","18025 Lake Iris Avenue","Baton Rouge","LA","70817","pammartin01@yahoo.com pammartin01@kw.com","(251)269-8864 or (251)279-0716","Are you ready for your vacation?  ","We hope you enjoy your stay. Please contact me If you need anything!"]
    page.append(companies)

    # Apply styles to the example sheet
    for i in range (1, len(TabTitles)+1): 
        print(TabTitles[i - 1])

        cell_obj = page.cell(row = 1, column = i)
        #cell_obj.value = TabTitles[i - 1]
        print(len(cell_obj.value))
        current_col = string.ascii_uppercase[i - 1]

        page.column_dimensions[current_col].width = len(cell_obj.value) * 1.5
        cell_obj.fill = tabColor_cell 
        cell_obj.font = tabColor_text
        print("") 

    # excel var utilities
    max_row = page.max_row
    max_col = page.max_column
    factor_charactersize = 1.18

    ##testing
    #print("testing")

    #var that store the max width of a cell in a column
    max_long_per_column = []

    #print("")

    #to fix the size of the cells in the temporal excel

    for x in range(0, max_col):
        current_col = string.ascii_uppercase[x]
        #numbers[x] = arr.array([0])
        #max_long_per_column[x].append(0)
        max_long_per_column.append(0)
        for y in range(1, max_row + 1):
            #save cell value
            cell_obj = page.cell(row = y, column = x + 1)
            print(cell_obj)
            print(type(cell_obj.value).__name__)

            if type(cell_obj.value).__name__ != 'NoneType':
                print(len(cell_obj.value))

                #
                if max_long_per_column[x] < len(cell_obj.value):
                    print("encontro uno mas grande")
                    max_long_per_column[x] = len(cell_obj.value)
                    print("el mas grande: " + str(max_long_per_column[x]))

                    page.column_dimensions[current_col].width = max_long_per_column[x] * factor_charactersize
                    pass
                else:
                    print("encontro uno mas chico")
                    pass
                pass
            else:
                print("WTF")
                pass
        print("Termino uno calumna --------------------------------------------")
        #check

        #page.column_dimensions[current_col].width = max_long_per_column[x] * 1.5
        
    #workbook_name = 'sample.xlsx'
    #workbook_name_temp = ''

    #excelFile = "{}\\~tempFile-{}.xlsx".format(sys.path[0], str(datetime.datetime.now()))
    #excelFile = "{}\\~tempfile.xlsx".format(sys.path[0])





    #print("Goto Sleep")
    #sleep(50)

    """
    for info in companies:
        page.append(info)
    """
    #wb.save(filename = workbook_name)

    wb.save(filename = tempfilename)


    #stuff_in_string = 'start excel.exe "{}"'.format(workbook_name)

    #stuff_in_string = 'start excel.exe "{}"'.format(excelFile)
    #stuff_in_string = 'start excel "{}"'.format(excelFile)

    #os.system(stuff_in_string)
    # return tempfilename
    pass


def excel_read(excel_file):
    import pandas
    excel_data_df = pandas.read_excel(excel_file, sheet_name=0)

    # print whole sheet data
    #print(excel_data_df)

    # get column names
    #print(excel_data_df.columns.ravel())

    #print(excel_data_df['First Name'].tolist())
    # https://www.journaldev.com/33306/pandas-read_excel-reading-excel-file-in-python
    #print('Excel Sheet to Dict:', excel_data_df.to_dict(orient='record'))
    #print('Excel Sheet to JSON:', excel_data_df.to_json(orient='records'))
    #print('Excel Sheet to CSV:\n', excel_data_df.to_csv(index=False))
    return json.loads(excel_data_df.to_json(orient='records'))


def getDataFromExcel(excel_file_to_load,subject,body,body_in_html):
    import os
    #excel_file_to_load = os.path.abspath(str(excel_file_to_load))

    file = excel_file_to_load
    wb = load_workbook(file, data_only=True)
    #ws = wb['TO_SERVICE-NOW']
    sheets = wb.sheetnames
    CurrentWorksheet = wb[sheets[0]]
    # CurrentWorksheet = wb.active

    max_row = CurrentWorksheet.max_row + 1
    max_col = CurrentWorksheet.max_column + 1


    for y in range(1, max_col):
        cell_Assignment_Group = CurrentWorksheet.cell(row=1, column=y).value 
        print("column_data_"+cell_Assignment_Group.replace(" ", "_").lower())
        #AssignedGroup = cell_Assignment_Group
        pass

    for x in range(2, max_row):
        #REQ_VAR = ""
        #RITM_VAR = ""
        #TASK_VAR = ""
        #GROUP_VALUE = ""

        #datetime.datetime.now()

        print("salto de linea -------------------------- " + time.strftime('%Y-%m-%d %H:%M:%S'))
        print("")

        # https://www.codespeedy.com/validate-email-in-python/
        """
        if CurrentWorksheet.cell(row=x, column=3).value == "NO":
        else:
            print("ya se hizo")
            print(CurrentWorksheet.cell(row=x, column=3).value)
            print("")
        """
        #print(ws.cell(row=x, column=y).value)
        column_data_first_name = CurrentWorksheet.cell(row=x, column=1).value 
        print(column_data_first_name)
        
        column_data_last_name = CurrentWorksheet.cell(row=x, column=2).value 
        print(column_data_last_name)
        
        column_data_email = CurrentWorksheet.cell(row=x, column=3).value 
        print(column_data_email)
        
        column_data_cell_phone = CurrentWorksheet.cell(row=x, column=4).value 
        print(column_data_cell_phone)
        
        column_data_unit = CurrentWorksheet.cell(row=x, column=5).value 
        print(column_data_unit)
        
        column_data_arrival = CurrentWorksheet.cell(row=x, column=6).value 
        column_data_arrival = getDateNormaliced(column_data_arrival)
        print(column_data_arrival)
        
        column_data_departure = CurrentWorksheet.cell(row=x, column=7).value 
        column_data_departure = getDateNormaliced(column_data_departure)
        print(column_data_departure)
        
        column_data_code = CurrentWorksheet.cell(row=x, column=8).value 
        print(column_data_code)
        
        column_data_company_name = CurrentWorksheet.cell(row=x, column=9).value 
        print(column_data_company_name)
        
        column_data_contact = CurrentWorksheet.cell(row=x, column=10).value 
        print(column_data_contact)
        
        column_data_maddress = CurrentWorksheet.cell(row=x, column=11).value 
        print(column_data_maddress)
        
        column_data_mcity = CurrentWorksheet.cell(row=x, column=12).value 
        print(column_data_mcity)
        
        column_data_mstate = CurrentWorksheet.cell(row=x, column=13).value 
        print(column_data_mstate)
        print("asda")
        
        column_data_mzip = CurrentWorksheet.cell(row=x, column=14).value 
        print(column_data_mzip)
        
        column_data_memail = CurrentWorksheet.cell(row=x, column=15).value 
        print(column_data_memail)
        
        column_data_phone = CurrentWorksheet.cell(row=x, column=16).value 
        print(column_data_phone)
        
        column_data_greeting = CurrentWorksheet.cell(row=x, column=17).value 
        print(column_data_greeting)
        
        column_data_salutation = CurrentWorksheet.cell(row=x, column=18).value 
        print(column_data_salutation)
        
        column_data_link_to_qr = CurrentWorksheet.cell(row=x, column=19).value 
        print(column_data_link_to_qr)

        #toServiceNow(cell_Assignment_Group.value , cell_Free_Text_Field.value)
        #CreateUserFreeformRequest("datavision","excuse me, this is a test",True)
        
        #CreateUserFreeformRequest(cell_Assignment_Group.value,cell_Free_Text_Field.value,True,x)
        
        
        #self.CreateUserFreeformRequest(cell_Assignment_Group,cell_Free_Text_Field,"NO")
        #self.CreateUserFreeformRequest()
        #mark as complete
        """
        CurrentWorksheet.cell(row=x, column=3).value = "YES"
        CurrentWorksheet.cell(row=x, column=4).value = self.request #REQUEST_VALUE
        CurrentWorksheet.cell(row=x, column=5).value = self.ritm # RITM_VALUE
        CurrentWorksheet.cell(row=x, column=6).value = self.task #TASK_VALUE
        CurrentWorksheet.cell(row=x, column=7).value = datetime.datetime.now()
        CurrentWorksheet.cell(row=x, column=8).value = self.AssignedGroupFromPage #ASSIGNMENT_GROUP_VALUE
        """

        """
        ws.cell(row=x, column=4).value = REQ_VAR #REQUEST_VALUE
        ws.cell(row=x, column=5).value = RITM_VAR # RITM_VALUE
        ws.cell(row=x, column=6).value = TASK_VAR #TASK_VALUE
        ws.cell(row=x, column=7).value = datetime.datetime.now()
        ws.cell(row=x, column=8).value = GROUP_VALUE #ASSIGNMENT_GROUP_VALUE
        """
        #CurrentWorksheet.cell(row=x, column=7).value = datetime.datetime.now()
        #print(datetime.datetime.now())
        #ws2['F5']
        # .splite(" ")
        qr_image_file = myQRCode.qrcode_create(column_data_code)

        # word_file_path = './~tempfile.docx'
        word_file_path = './document.docx'
        qr_image_file = myConfig.getPath(qr_image_file)

        myWord.word_create_from_template(word_file_path,qr_image_file,str(column_data_first_name) +" "+ str(column_data_last_name),column_data_greeting,column_data_unit,column_data_arrival,column_data_departure,column_data_salutation,column_data_maddress,column_data_company_name,column_data_phone)

        body_from_template = myWord.create_email_from_html_template(qr_image_file,str(column_data_first_name) +" "+ str(column_data_last_name),column_data_greeting,column_data_unit,column_data_arrival,column_data_departure,column_data_salutation,column_data_maddress,column_data_company_name,column_data_phone)

        print(type(column_data_email))
        if type(column_data_email) == str:
            import re
            listOfEmails = re.split('\s+', column_data_email)
            print(listOfEmails)
            try:
                print(column_data_memail.split(" "))
                pass
            except AttributeError as NoneType:
                pass
            
            for email in listOfEmails:

                print("sending email to : "+email)
                # myMail.email_send("lucasain2010@gmail.com","test subject","test body body","<p>my test html body</p>",wordfile)
                #myMail.email_send(email,"Hello World","test body body","<p>my test html body</p>",word_file_path)
                
                #print(email,subject,body,body_in_html,word_file_path)
                #myMail.email_send(email,subject,body,body_in_html,word_file_path)
                # send the email with the image
                if myConfig.check_data('Automate outlook mailing','Word_File_Attachement'):
                    if myConfig.get_saved_data('Automate outlook mailing','Word_File_Attachement') == "True":
                        
                        myMail.email_send(email,subject,body,body_from_template,qr_image_file,word_file_path)
                        print('Word File Attachement On')
                        pass
                    else:
                        print('Word File Attachement Off')
                        myMail.email_send(email,subject,body,body_from_template,qr_image_file)
                        pass
                    pass
                else:
                    print('auto_run not exist')
                    myConfig.configuration_file_create_persist()
                    pass
                pass
                
                #,subject,body,body_in_html
                pass
            pass
        else:
            print("not email allowed")
            pass
        print("")
        print("mark as complete")
        print("")
        #ws['A10'] = datetime.datetime.now()
        #save escel
        #text_temp = "successfully_excel_"
        
        
        #sleep(15)
        pass

    wb.save(file)
    #sleep(120)
    #driver.close()
    #https://stackoverflow.com/questions/21191494/how-to-open-an-excel-file-with-python-to-display-its-content
    #https://www.w3schools.com/python/python_file_remove.asp
    pass


# getDataFromExcel("~tempfile.1.xlsx")